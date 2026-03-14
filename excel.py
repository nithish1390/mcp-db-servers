"""
MCP server for Excel operations including charts, pivot tables, and data analysis.
"""

import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart
from openpyxl.chart import Reference, Series
from openpyxl.styles import Font, PatternFill
import tempfile
import os
from collections.abc import AsyncIterator
from contextlib import asynccontextmanager
from dataclasses import dataclass
from typing import List, Dict, Any

from mcp.server.fastmcp import Context, FastMCP


@dataclass
class ExcelContext:
    """Context for Excel operations."""
    temp_dir: str


@asynccontextmanager
async def excel_lifespan(server: FastMCP) -> AsyncIterator[ExcelContext]:
    """Manage Excel server lifecycle."""
    temp_dir = tempfile.mkdtemp(prefix="excel_mcp_")
    try:
        yield ExcelContext(temp_dir=temp_dir)
    finally:
        # Clean up temp files
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)


# Create FastMCP server
mcp = FastMCP("Excel Operations Server", lifespan=excel_lifespan)


@mcp.tool()
def create_excel_file(filename: str, data: str, ctx: Context) -> str:
    """Create a new Excel file with data.

    Args:
        filename: Name of the Excel file to create
        data: CSV formatted data (comma-separated values)

    Returns:
        Success message with file path
    """
    try:
        # Parse CSV data
        lines = data.strip().split('\n')
        if not lines:
            return "Error: No data provided"

        # Create DataFrame
        headers = lines[0].split(',')
        rows = [line.split(',') for line in lines[1:]]
        df = pd.DataFrame(rows, columns=headers)

        # Save to Excel
        temp_dir = ctx.request_context.lifespan_context.temp_dir
        filepath = os.path.join(temp_dir, f"{filename}.xlsx")
        df.to_excel(filepath, index=False)

        return f"Excel file created successfully: {filepath}"
    except Exception as e:
        return f"Error creating Excel file: {e}"


@mcp.tool()
def create_chart(ctx: Context, data: str, chart_type: str, x_column: str, y_column: str,
                title: str = "Chart", filename: str = "chart") -> str:
    """Create a chart from data and save as Excel file with embedded chart.

    Args:
        data: CSV formatted data
        chart_type: Type of chart (bar, line, pie, scatter)
        x_column: Column name for X-axis
        y_column: Column name for Y-axis
        title: Chart title
        filename: Output filename

    Returns:
        Success message with file path
    """
    try:
        # Parse data
        lines = data.strip().split('\n')
        headers = lines[0].split(',')
        rows = [line.split(',') for line in lines[1:]]
        df = pd.DataFrame(rows, columns=headers)

        # Convert numeric columns
        for col in [x_column, y_column]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        # Create Excel file with chart
        temp_dir = ctx.request_context.lifespan_context.temp_dir
        filepath = os.path.join(temp_dir, f"{filename}.xlsx")

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Data', index=False)

            # Create chart sheet
            workbook = writer.book
            chart_sheet = workbook.create_sheet('Chart')

            # Create chart based on type
            if chart_type.lower() == 'bar':
                chart = BarChart()
            elif chart_type.lower() == 'line':
                chart = LineChart()
            elif chart_type.lower() == 'pie':
                chart = PieChart()
            elif chart_type.lower() == 'scatter':
                chart = ScatterChart()
            else:
                chart = BarChart()

            # Add data to chart
            x_col_idx = headers.index(x_column) + 1
            y_col_idx = headers.index(y_column) + 1

            x_values = Reference(writer.sheets['Data'], min_col=x_col_idx, min_row=2, max_row=len(df)+1)
            y_values = Reference(writer.sheets['Data'], min_col=y_col_idx, min_row=2, max_row=len(df)+1)

            series = Series(y_values, x_values, title=y_column)
            chart.append(series)

            chart.title = title
            chart.x_axis.title = x_column
            chart.y_axis.title = y_column

            # Add chart to sheet
            chart_sheet.add_chart(chart, "A1")

        return f"Chart created successfully: {filepath}"
    except Exception as e:
        return f"Error creating chart: {e}"


@mcp.tool()
def create_pivot_table(ctx: Context, data: str, values: str, index: str, columns: str = "",
                      aggfunc: str = "sum", filename: str = "pivot") -> str:
    """Create a pivot table from data.

    Args:
        data: CSV formatted data
        values: Column to aggregate
        index: Column for row labels
        columns: Column for column labels (optional)
        aggfunc: Aggregation function (sum, mean, count, etc.)
        filename: Output filename

    Returns:
        Success message with file path
    """
    try:
        # Parse data
        lines = data.strip().split('\n')
        headers = lines[0].split(',')
        rows = [line.split(',') for line in lines[1:]]
        df = pd.DataFrame(rows, columns=headers)

        # Convert numeric columns
        df = df.apply(pd.to_numeric, errors='ignore')

        # Create pivot table
        pivot_args = {
            'values': values,
            'index': index,
            'aggfunc': aggfunc
        }
        if columns:
            pivot_args['columns'] = columns

        pivot_df = pd.pivot_table(df, **pivot_args)

        # Save to Excel
        temp_dir = ctx.request_context.lifespan_context.temp_dir
        filepath = os.path.join(temp_dir, f"{filename}.xlsx")

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Original data
            df.to_excel(writer, sheet_name='Data', index=False)
            # Pivot table
            pivot_df.to_excel(writer, sheet_name='Pivot', index=True)

        return f"Pivot table created successfully: {filepath}"
    except Exception as e:
        return f"Error creating pivot table: {e}"


@mcp.tool()
def analyze_data(data: str, analysis_type: str, ctx: Context) -> str:
    """Perform data analysis on the provided data.

    Args:
        data: CSV formatted data
        analysis_type: Type of analysis (summary, correlation, distribution)

    Returns:
        Analysis results
    """
    try:
        # Parse data
        lines = data.strip().split('\n')
        headers = lines[0].split(',')
        rows = [line.split(',') for line in lines[1:]]
        df = pd.DataFrame(rows, columns=headers)

        # Convert numeric columns
        numeric_df = df.apply(pd.to_numeric, errors='ignore')

        if analysis_type.lower() == 'summary':
            # Statistical summary
            result = numeric_df.describe().to_string()

        elif analysis_type.lower() == 'correlation':
            # Correlation matrix
            numeric_cols = numeric_df.select_dtypes(include=['number']).columns
            if len(numeric_cols) < 2:
                result = "Need at least 2 numeric columns for correlation analysis"
            else:
                corr = numeric_df[numeric_cols].corr()
                result = corr.to_string()

        elif analysis_type.lower() == 'distribution':
            # Distribution analysis
            result = ""
            for col in numeric_df.select_dtypes(include=['number']).columns:
                result += f"\n{col} Distribution:\n"
                result += f"Mean: {numeric_df[col].mean():.2f}\n"
                result += f"Median: {numeric_df[col].median():.2f}\n"
                result += f"Std Dev: {numeric_df[col].std():.2f}\n"
                result += f"Min: {numeric_df[col].min()}\n"
                result += f"Max: {numeric_df[col].max()}\n"

        else:
            result = f"Unknown analysis type: {analysis_type}. Available: summary, correlation, distribution"

        return result
    except Exception as e:
        return f"Error performing analysis: {e}"


@mcp.tool()
def format_excel_file(filepath: str, formatting: str, ctx: Context) -> str:
    """Apply formatting to an existing Excel file.

    Args:
        filepath: Path to the Excel file
        formatting: Formatting options (header_bold, alternate_rows, borders)

    Returns:
        Success message
    """
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active

        if 'header_bold' in formatting:
            # Make header row bold
            for cell in sheet[1]:
                cell.font = Font(bold=True)

        if 'alternate_rows' in formatting:
            # Alternate row colors
            for row in range(2, sheet.max_row + 1):
                if row % 2 == 0:
                    for cell in sheet[row]:
                        cell.fill = PatternFill(start_color="FFE6E6FA", end_color="FFE6E6FA", fill_type="solid")

        if 'borders' in formatting:
            from openpyxl.styles import Border, Side
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

        workbook.save(filepath)
        return f"Formatting applied successfully to {filepath}"
    except Exception as e:
        return f"Error applying formatting: {e}"


@mcp.tool()
def create_excel(ctx: Context, filename: str, sheets_data: str, sheet_names: str = "Sheet1") -> str:
    """Create a dynamic Excel file with multiple sheets and various data formats.

    Args:
        filename: Name of the Excel file to create
        sheets_data: JSON-like string with sheet data. Format: {"sheet1": [["row1col1", "row1col2"], ["row2col1", "row2col2"]], "sheet2": [["data"]]}
        sheet_names: Comma-separated list of sheet names (optional, will use Sheet1, Sheet2, etc. if not provided)

    Returns:
        Success message with file path
    """
    try:
        import json
        import ast

        # Parse sheets_data - try JSON first, then literal eval
        try:
            data_dict = json.loads(sheets_data)
        except json.JSONDecodeError:
            try:
                data_dict = ast.literal_eval(sheets_data)
            except:
                return "Error: Invalid data format. Please provide data as a valid dictionary string."

        # Parse sheet names
        names = [name.strip() for name in sheet_names.split(',')] if sheet_names else []

        # Create workbook
        workbook = openpyxl.Workbook()

        # Remove default sheet
        workbook.remove(workbook.active)

        # Create sheets and populate data
        for i, (sheet_key, sheet_data) in enumerate(data_dict.items()):
            sheet_name = names[i] if i < len(names) else f"Sheet{i+1}"
            sheet = workbook.create_sheet(sheet_name)

            if isinstance(sheet_data, list):
                for row_idx, row in enumerate(sheet_data, 1):
                    if isinstance(row, list):
                        for col_idx, cell_value in enumerate(row, 1):
                            sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    else:
                        # Single value in row
                        sheet.cell(row=row_idx, column=1, value=row)
            else:
                # Single value
                sheet.cell(row=1, column=1, value=sheet_data)

        # Save file
        temp_dir = ctx.request_context.lifespan_context.temp_dir
        filepath = os.path.join(temp_dir, f"{filename}.xlsx")
        workbook.save(filepath)

        return f"Dynamic Excel file created successfully: {filepath}"
    except Exception as e:
        return f"Error creating dynamic Excel file: {e}"


@mcp.tool()
def update_excel(ctx: Context, filepath: str, updates: str) -> str:
    """Update an existing Excel file with new data.

    Args:
        filepath: Path to the existing Excel file
        updates: JSON-like string with update operations. Format: [{"sheet": "Sheet1", "cell": "A1", "value": "new_value"}, {"sheet": "Sheet1", "range": "B2:C3", "data": [["val1", "val2"], ["val3", "val4"]]}]

    Returns:
        Success message
    """
    try:
        import json
        import ast
        from openpyxl.utils import column_index_from_string, get_column_letter

        # Parse updates - try JSON first, then literal eval
        try:
            update_list = json.loads(updates)
        except json.JSONDecodeError:
            try:
                update_list = ast.literal_eval(updates)
            except:
                return "Error: Invalid updates format. Please provide updates as a valid list of dictionaries."

        # Load workbook
        if not os.path.exists(filepath):
            return f"Error: File {filepath} does not exist"

        workbook = openpyxl.load_workbook(filepath)

        for update in update_list:
            sheet_name = update.get('sheet', 'Sheet1')

            # Get or create sheet
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(sheet_name)

            if 'cell' in update and 'value' in update:
                # Update single cell
                cell_ref = update['cell']
                value = update['value']
                sheet[cell_ref] = value

            elif 'range' in update and 'data' in update:
                # Update range of cells
                range_ref = update['range']
                data = update['data']

                # Parse range (e.g., "A1:B2")
                start_cell, end_cell = range_ref.split(':')
                start_col = column_index_from_string(start_cell[0])
                start_row = int(start_cell[1:])
                end_col = column_index_from_string(end_cell[0])
                end_row = int(end_cell[1:])

                # Fill data
                for row_idx, row_data in enumerate(data):
                    if isinstance(row_data, list):
                        for col_idx, cell_value in enumerate(row_data):
                            if start_row + row_idx <= end_row and start_col + col_idx <= end_col:
                                sheet.cell(row=start_row + row_idx, column=start_col + col_idx, value=cell_value)
                    else:
                        # Single value in row
                        if start_row + row_idx <= end_row:
                            sheet.cell(row=start_row + row_idx, column=start_col, value=row_data)

            elif 'append_row' in update and 'data' in update:
                # Append row to sheet
                data = update['data']
                next_row = sheet.max_row + 1

                if isinstance(data, list):
                    for col_idx, cell_value in enumerate(data, 1):
                        sheet.cell(row=next_row, column=col_idx, value=cell_value)
                else:
                    sheet.cell(row=next_row, column=1, value=data)

            elif 'append_column' in update and 'data' in update:
                # Append column to sheet
                data = update['data']
                next_col = sheet.max_column + 1

                if isinstance(data, list):
                    for row_idx, cell_value in enumerate(data, 1):
                        sheet.cell(row=row_idx, column=next_col, value=cell_value)
                else:
                    sheet.cell(row=1, column=next_col, value=data)

        # Save updated file
        workbook.save(filepath)
        return f"Excel file updated successfully: {filepath}"
    except Exception as e:
        return f"Error updating Excel file: {e}"


@mcp.resource("excel://help")
def get_excel_help() -> str:
    """Get help information for Excel operations."""
    return """
Excel Operations Available:

Tools:
- create_excel_file: Create Excel file from CSV data
- create_excel: Create dynamic Excel file with multiple sheets and flexible data
- update_excel: Update existing Excel files with new data
- create_chart: Create charts (bar, line, pie, scatter)
- create_pivot_table: Create pivot tables with aggregation
- analyze_data: Perform data analysis (summary, correlation, distribution)
- format_excel_file: Apply formatting to Excel files

Chart Types: bar, line, pie, scatter
Analysis Types: summary, correlation, distribution
Formatting Options: header_bold, alternate_rows, borders

create_excel Data Format:
{"Sheet1": [["Name", "Age"], ["Alice", 30], ["Bob", 25]], "Sheet2": [["Data"]]}

update_excel Operations Format:
[{"sheet": "Sheet1", "cell": "A1", "value": "New Value"},
 {"sheet": "Sheet1", "range": "B2:C3", "data": [["val1", "val2"], ["val3", "val4"]]},
 {"sheet": "Sheet1", "append_row": true, "data": ["New", "Row", "Data"]},
 {"sheet": "Sheet1", "append_column": true, "data": ["New", "Column", "Data"]}]

Example CSV Data Format:
Name,Age,Salary
Alice,30,50000
Bob,25,45000
Charlie,35,60000
    """


if __name__ == "__main__":
    mcp.run()